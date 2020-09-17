import cgitb
cgitb.enable()

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

workbook = load_workbook(filename="Foods.xlsx")
workbook.sheetnames
soups_sheet = workbook["Soups"]
dishes_sheet = workbook["Dishes"]
sides_sheet = workbook["Sides"]
entree_sheet = workbook["Entree"]
snack_sheet = workbook["Snack"]
desserts_sheet = workbook["Desserts"]
breakfast_sheet = workbook["Breakfast"]

import random

while True:
    type=str(input("Please choose one option: Soups or Dishes or  Entrees with sides or Snacks or Desserts or Breakfast. \nIf you want to exit the program type EXIT. \nIf you wanna add a new food to the list typ ADD. :  "))
    if type.lower() in ["soup", "soups"]:
        total_rows = soups_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p="A"+str(s)
        t = p[0] + p[2]
        print(soups_sheet[t].value)
        continue
    elif type.lower() in ["dish", "dishes"]:
        total_rows = dishes_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p = "A" + str(s)
        t = p[0] + p[2]
        print(dishes_sheet[t].value)
        continue
    elif type.lower() in["entree", "entrees", "side", "sides", "entrees with sides", "entree with side", "entrees with side", "entree with sides"]:
        total_rows = sides_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p = "A" + str(s)
        t = p[0] + p[2]
        total_rows = entree_sheet.max_row
        c = random.sample(range(2, total_rows), k=1)
        d = "A" + str(c)
        e = d[0] + d[2]
        print(sides_sheet[t].value+" with "+entree_sheet[e].value)
        continue
    elif type.lower() in["snack", "snacks"]:
        total_rows = snack_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p = "A" + str(s)
        t = p[0] + p[2]
        print(snack_sheet[t].value)
        continue
    elif type.lower() in["dessert", "desserts"]:
        total_rows = desserts_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p = "A" + str(s)
        t = p[0] + p[2]
        print(desserts_sheet[t].value)
        continue
    elif type.lower() in ["breakfast", "breakfasts"]:
        total_rows = breakfast_sheet.max_row
        s = random.sample(range(2, total_rows), k=1)
        p = "A" + str(s)
        t = p[0] + p[2]
        print(breakfast_sheet[t].value)
        continue
    elif type in ["add"]:
        add = str(input(
               "Please choose what type of food do you want to add to the list: Soups or Dishes or Entrees or Side or Snacks or Desserts or Breakfast: "))
        if add.lower() in ["soup", "soups"]:
               food = str(input("Please type in what kind of soup do you want to add to the list: "))
               total_rows = soups_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               soups_sheet[p]= food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["dish", "dishes"]:
               food = str(input("Please type in what kind of dish do you want to add to the list: "))
               total_rows = dishes_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               dishes_sheet[p]= food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["entree", "entrees"]:
               food = str(input("Please type in what kind of entree do you want to add to the list: "))
               total_rows = entree_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               entree_sheet[p] = food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["side", "sides"]:
               food = str(input("Please type in what kind of side do you want to add to the list: "))
               total_rows = sides_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               sides_sheet[p] = food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["snack", "snacks"]:
               food = str(input("Please type in what kind of snack do you want to add to the list: "))
               total_rows = snack_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               snack_sheet[p] = food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["dessert", "desserts"]:
               food = str(input("Please type in what kind of dessert do you want to add to the list: "))
               total_rows = desserts_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               desserts_sheet[p] = food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        elif add.lower() in ["breakfast", "breakfasts"]:
               food = str(input("Please type in what kind of breakfast do you want to add to the list: "))
               total_rows = breakfast_sheet.max_row
               next_row = total_rows + 1
               p = "A" + str(next_row)
               breakfast_sheet[p] = food
               workbook.save('foods.xlsx')
               print("The expansion was successful.")
        else:
               print("Please choose an option from the specified list.")
        continue
    elif type.lower() in ["exit"]:
        break
    else:
        print("Please choose an option from the specified list.")
        continue
